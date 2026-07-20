import os
import sys
import json
import time
import openpyxl
from typing import List, Optional, Dict, Any
from pydantic import BaseModel, Field
from google import genai
from google.genai import types
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# Safe path config
project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if project_root not in sys.path:
    sys.path.insert(0, project_root)

from core.agent import load_unified_config

def call_gemini_with_retry(client, model, contents, config=None, max_attempts=5):
    import re
    import time
    
    # Introduce rate limiting sleep to avoid hitting immediate RPM limits
    time.sleep(5.0)
    
    for attempt in range(1, max_attempts + 1):
        try:
            return client.models.generate_content(
                model=model,
                contents=contents,
                config=config
            )
        except Exception as ex:
            ex_str = str(ex)
            is_rate_limit = any(term in ex_str for term in ["429", "RESOURCE_EXHAUSTED", "Quota exceeded"])
            is_unavailable = any(term in ex_str for term in ["503", "UNAVAILABLE"])
            
            if is_rate_limit or is_unavailable:
                if is_rate_limit:
                    match = re.search(r'(?:retry in|retryDelay[\'\"\s:]+)([\d\.]+)', ex_str)
                    if match:
                        sleep_time = float(match.group(1)) + 5.0
                    else:
                        sleep_time = (2 ** attempt) * 10
                    alert_msg = f"Gemini API 429 Exhausted"
                else:
                    sleep_time = 60.0
                    alert_msg = f"Gemini API 503 Unavailable"
                
                print(f"[RATE LIMIT ALERT] {alert_msg}. Automatically sleeping for {sleep_time:.2f} seconds before retrying execution loop (Attempt {attempt}/{max_attempts})...")
                time.sleep(sleep_time)
            else:
                raise ex
    raise Exception(f"Gemini API Rate Limit or Availability attempts exhausted ({max_attempts} attempts).")


def safe_save_workbook(wb, excel_path: str):
    """Saves workbook safely. On file-locking PermissionError, writes to a timestamped fallback recovery file."""
    import time
    if not os.path.isabs(excel_path):
        excel_path = os.path.join(project_root, excel_path)
    os.makedirs(os.path.dirname(excel_path), exist_ok=True)
    try:
        wb.save(excel_path)
    except (PermissionError, OSError, IOError) as lock_err:
        timestamp = int(time.time())
        dirname = os.path.dirname(excel_path)
        basename = os.path.splitext(os.path.basename(excel_path))[0]
        fallback_path = os.path.join(dirname, f"{basename}_fallback_{timestamp}.xlsx")
        print(f"⚠️ [EXCEL LOCK WARNING]: Primary workbook '{excel_path}' is locked or open in background ({lock_err}). Writing recovery backup to '{fallback_path}'.")
        try:
            wb.save(fallback_path)
            print(f"💾 [RECOVERY BACKUP SAVED]: Telemetry and matrix saved cleanly to {fallback_path}")
        except Exception as e_fb:
            print(f"❌ Failed to save fallback workbook: {e_fb}")


def write_pipeline_failure_to_excel(excel_path: str, error_message: str):
    import os
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    
    os.makedirs(os.path.dirname(excel_path), exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Suite Matrix"
    ws.views.sheetView[0].showGridLines = True
    
    headers = ["Test Case ID", "Component", "Description", "Execution Goal", "Expected Result", "Status", "Target URL", "Timestamp", "Screenshot Link", "Issue Type"]
    ws.append(headers)
    
    ws.append([
        "TC000_ERR",
        "Pipeline",
        "Pipeline execution failed during ingestion/generation phase.",
        f"Error: {error_message}",
        "Pipeline should complete successfully.",
        "FAILED",
        "N/A",
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "",
        "N/A"
    ])
    
    fail_fill = PatternFill(start_color="FCE5CD", end_color="FCE5CD", fill_type="solid")
    fail_font = Font(name="Segoe UI", size=10, bold=True, color="A61C00")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=2, column=col)
        cell.font = fail_font
        if col == 6:
            cell.fill = fail_fill
            
    safe_save_workbook(wb, excel_path)
    print(f"💾 Saved pipeline failure details to {excel_path}")


class JiraStoryData(BaseModel):
    title: str = Field(description="The summary or title of the Jira story/ticket.")
    description: str = Field(description="The main description body of the story.")
    acceptance_criteria: List[str] = Field(description="List of acceptance criteria extracted from the story.")
    target_url: Optional[str] = Field(None, description="The test environment web application URL if mentioned in the story description.")
    issue_type: Optional[str] = Field("User Story", description="The issue type of the ticket, e.g., User Story, Bug.")


class TestCaseItem(BaseModel):
    test_case_id: str = Field(description="Unique identifier, e.g. TC001, TC002.")
    component: str = Field(description="The component area, e.g. Login, Authentication, Form submission, Validation.")
    description: str = Field(description="Detailed explanation of what is being tested.")
    execution_goal: str = Field(description="Specific, detailed natural language instructions telling our autonomous web agent exactly what to do step-by-step.")
    expected_result: str = Field(description="Expected outcome or success indicator to check.")


class TestSuiteOutput(BaseModel):
    test_cases: List[TestCaseItem] = Field(description="The complete list of predicted test case scenarios.")


def generate_dynamic_synthetic_story() -> JiraStoryData:
    import random
    from faker import Faker
    fake = Faker()
    
    target_url = "https://dev.urbuddi.com/login"
    try:
        config = load_unified_config()
        target_url = config["environment"].get("target_url") or config["environment"].get("default_url") or target_url
    except Exception:
        pass

    actions = [
        "Authenticate and access dashboard",
        "Submit transaction form",
        "Update user profile settings",
        "Apply monthly leave requests",
        "Process payroll calculations",
        "Generate analytics report summary",
        "Configure notification preferences"
    ]
    
    rules = [
        "must validate all mandatory input fields and reject empty submissions",
        "should handle boundary conditions and enforce numeric limits",
        "must trigger security alerts on multiple validation failures",
        "must show success notifications and redirect to active view URL",
        "should calculate zero-state values correctly if inputs are missing"
    ]
    
    triggers = [
        "under high system load parameters",
        "with dynamic synthetic credentials",
        "within the configured session timeout period",
        "across all responsive viewport configurations"
    ]
    
    title_action = random.choice(actions)
    title = f"Dynamic Verification: {title_action} on Target Application"
    
    description = (
        f"As a QA verification system, I want to programmatically execute: {title_action.lower()} "
        f"so that I can validate that the system behaves correctly under rules such as: '{random.choice(rules)}'. "
        f"This test is run using randomly synthesized inputs {random.choice(triggers)}."
    )
    
    num_ac = random.randint(3, 5)
    ac_list = []
    for i in range(num_ac):
        ac_list.append(
            f"Verify that performing action '{random.choice(actions).lower()}' "
            f"conforms successfully to standard requirement: '{random.choice(rules)}'."
        )
        
    return JiraStoryData(
        title=title,
        description=description,
        acceptance_criteria=ac_list,
        target_url=target_url,
        issue_type="User Story"
    )


class TestCaseGenerator:
    """Generates test cases from Jira criteria and writes to Excel workbook."""
    
    def __init__(self, api_key: str):
        self.api_key = api_key
        self.client = genai.Client(api_key=api_key)

    def generate_suite(self, story: JiraStoryData, target_url: Optional[str] = None) -> List[TestCaseItem]:
        """Uses Gemini to predict scenarios based on story criteria."""
        print("🧠 TestCaseGenerator: Invoking Gemini to generate test cases...")
        
        url_prompt_addition = ""
        if target_url:
            url_prompt_addition = f"\nAll test cases must be designed to be executed starting from this baseline URL: {target_url}\n"
            
        try:
            config = load_unified_config()
            test_data = config.get("test_data", {})
            config_user = test_data.get("username", "")
            config_pass = test_data.get("password", "")
            if config_user and config_pass:
                url_prompt_addition += f"\nFor any authentication scenario, prioritize using these credentials: Username: {config_user}, Password: {config_pass}. Do not generate synthetic placeholders.\n"
        except Exception:
            pass
            
        issue_type = story.issue_type or "User Story"
        if issue_type.lower() == "bug":
            type_prompt_addition = (
                "\nThis issue is identified as a 'Bug'. Prioritize targeted edge-case verification, "
                "regression tests, and validation of the fix to ensure the defect is resolved.\n"
            )
        else:
            type_prompt_addition = (
                "\nThis issue is identified as a 'User Story'. Focus on functional acceptance criteria "
                "and happy path validation.\n"
            )

        prompt = (
            "You are a Senior Principal QA Engineer. Analyse the following Jira User Story details:\n"
            f"Title: {story.title}\n"
            f"Description: {story.description}\n"
            f"Acceptance Criteria:\n" + "\n".join(f"- {ac}" for ac in story.acceptance_criteria) + "\n\n"
            f"{url_prompt_addition}"
            f"{type_prompt_addition}"
            "Predict and generate a comprehensive suite of logical test case scenarios.\n"
            "CRITICAL SCOPING & STRUCTURE REQUIREMENTS:\n"
            "- Focus strictly on the functional target of this user story (e.g., Leave Allocation and Salary Disbursal rules, calculations, allocations, etc.).\n"
            "- Do NOT generate boilerplate or generic authentication/login validation scenarios (such as checking invalid passwords, empty fields, etc., often labeled generic login checks TC001-TC006). Assume baseline authentication is handled.\n"
            "- Do NOT write a single long, combined sequence of all scenarios in one test case.\n"
            "- You MUST dynamically evaluate the requirements to generate distinct, atomic test cases covering:\n"
            "  1. Happy Path Validation (e.g., perfect full-month leave processing or clean success flow).\n"
            "  2. Boundary Conditions (e.g., leave crossing over two payroll cycles, limits, thresholds).\n"
            "  3. Negative/Edge Validations (e.g., partial leaves, zero balances, retro-active cancellations, malformed parameters).\n\n"
            "Return the list of test cases matching the structured schema."
        )

        response = call_gemini_with_retry(
            client=self.client,
            model='gemini-2.5-flash',
            contents=prompt,
            config=types.GenerateContentConfig(
                response_mime_type="application/json",
                response_schema=TestSuiteOutput,
            )
        )
        
        suite = json.loads(response.text.strip())
        cases = [TestCaseItem(**c) for c in suite.get("test_cases", [])]
        print(f"📊 TestCaseGenerator: Generated {len(cases)} test cases.")
        return cases

    def write_to_excel(self, cases: List[TestCaseItem], default_url: str, output_path: str = "outputs/test_suite.xlsx", issue_type: str = "User Story"):
        """Saves generated test cases into a highly polished, formatted Excel workbook."""
        if not os.path.isabs(output_path):
            output_path = os.path.join(project_root, output_path)
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Test Suite Matrix"
        
        # Enable grid lines
        ws.views.sheetView[0].showGridLines = True
        
        headers = [
            "Test Case ID", "Component", "Description", "Execution Goal", "Expected Result",
            "Status", "Target URL", "Timestamp", "Screenshot Link", "Issue Type"
        ]
        ws.append(headers)
        
        # Format Headers
        header_fill = PatternFill(start_color="3B4F67", end_color="3B4F67", fill_type="solid")
        header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        border_side = Side(border_style="thin", color="D3D3D3")
        thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
        
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
            
        data_font = Font(name="Segoe UI", size=10, color="333333")
        pending_fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")
        pending_font = Font(name="Segoe UI", size=10, bold=True, color="666666")
        
        for idx, case in enumerate(cases):
            row_idx = idx + 2
            ws.append([
                case.test_case_id,
                case.component,
                case.description,
                case.execution_goal,
                case.expected_result,
                "PENDING",
                default_url,
                "",
                "",
                issue_type
            ])
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col)
                cell.font = data_font
                cell.border = thin_border
                
                if col in [1, 2, 6, 7, 8, 9, 10]:
                    cell.alignment = center_align
                else:
                    cell.alignment = left_align
                    
                if col == 6:  # Status
                    cell.fill = pending_fill
                    cell.font = pending_font
                    
        # Auto-adjust column widths
        for col in ws.columns:
            max_len = 0
            for cell in col:
                val = str(cell.value or "")
                if len(val) > max_len:
                    max_len = len(val)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 40)
            
        ws.row_dimensions[1].height = 28
        for row in range(2, len(cases) + 2):
            ws.row_dimensions[row].height = 24
            
        safe_save_workbook(wb, output_path)
        print(f"💾 TestCaseGenerator: Saved workbook to {output_path}")
